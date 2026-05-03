"""ETL pipeline: parse Excel, classify, compute risk & next action."""
from __future__ import annotations
import csv
import io
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional

from openpyxl import load_workbook

from models import Statement, Category, Status, RiskLevel


# Rupee-value parsing
def parse_rupee_value(val: Any) -> float:
    """Parse a ₹ value (may contain commas, Cr, Lakh) and return value in Crore."""
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        # Raw rupee amount — convert to crore (1 Cr = 10,000,000)
        n = float(val)
        if n <= 0:
            return 0.0
        # If already < 10000 we assume it is already in Cr; else raw ₹
        return round(n / 10_000_000, 4) if n > 10_000 else round(n, 4)
    s = str(val).strip()
    if not s:
        return 0.0
    s_lower = s.lower()
    # Clean
    cleaned = re.sub(r"[₹,\s]", "", s)
    cleaned = cleaned.replace("nil", "").replace("NIL", "")
    if not cleaned:
        return 0.0
    # Extract numeric
    try:
        # Some cells use "Cr", "Lakh"
        mult = 1.0
        if "cr" in s_lower:
            cleaned = re.sub(r"(?i)cr", "", cleaned)
            mult = 10_000_000  # pretend it's in rupees with cr hint — override below
            # Already in Cr, so we want to just get the number
            n = float(re.sub(r"[^0-9.\-]", "", cleaned) or 0)
            return round(n, 4)
        if "lakh" in s_lower or s_lower.endswith("l"):
            cleaned = re.sub(r"(?i)lakh|l$", "", cleaned)
            n = float(re.sub(r"[^0-9.\-]", "", cleaned) or 0)
            return round(n / 100.0, 4)  # 1 Cr = 100 Lakh
        n = float(re.sub(r"[^0-9.\-]", "", cleaned) or 0)
        if n <= 0:
            return 0.0
        # Assume raw rupees if big number
        return round(n / 10_000_000, 4) if n > 10_000 else round(n, 4)
    except (ValueError, TypeError):
        return 0.0


# Map sheet name to (statement, category) — includes common spelling variants
_SHEET_ENTRIES = [
    ("A- Medicine", Statement.A, Category.MEDICINE),
    ("A- Medicine ", Statement.A, Category.MEDICINE),
    ("A- Equipments", Statement.A, Category.EQUIPMENT),
    ("A- Equipment", Statement.A, Category.EQUIPMENT),
    ("B- Medicine", Statement.B, Category.MEDICINE),
    ("B- Equipments", Statement.B, Category.EQUIPMENT),
    ("B- Equipment", Statement.B, Category.EQUIPMENT),
    ("C- Medicine", Statement.C, Category.MEDICINE),
    ("C- Equipments", Statement.C, Category.EQUIPMENT),
    ("C- Equipment", Statement.C, Category.EQUIPMENT),
    ("D-Medicine", Statement.D, Category.MEDICINE),
    ("D- Medicine", Statement.D, Category.MEDICINE),
    ("D-Equipments", Statement.D, Category.EQUIPMENT),
    ("D- Equipments", Statement.D, Category.EQUIPMENT),
    ("D- Equipment", Statement.D, Category.EQUIPMENT),
]


def _normalize_sheet_title(name: str) -> str:
    n = name.strip().lower()
    n = n.replace("–", "-").replace("—", "-")
    n = re.sub(r"\s+", " ", n)
    n = re.sub(r"\s*-\s*", "-", n)
    return n


def _build_sheet_lookup() -> Dict[str, Tuple[Statement, Category]]:
    out: Dict[str, Tuple[Statement, Category]] = {}
    for raw, stmt, cat in _SHEET_ENTRIES:
        out[raw] = (stmt, cat)
        out[raw.strip()] = (stmt, cat)
        out[_normalize_sheet_title(raw)] = (stmt, cat)
    return out


SHEET_MAP = _build_sheet_lookup()


def _lookup_sheet_name(sname: str) -> Optional[Tuple[Statement, Category]]:
    if not sname or not str(sname).strip():
        return None
    s = str(sname).strip()
    if s in SHEET_MAP:
        return SHEET_MAP[s]
    key = _normalize_sheet_title(s)
    return SHEET_MAP.get(key)


def _infer_stmt_cat_from_filename(stem: str) -> Optional[Tuple[Statement, Category]]:
    """
    Single-sheet workbooks like report_a.xlsx / report_b.xlsx often use generic tab names.
    Infer A/B/C/D from the file stem; category defaults to Medicine unless filename hints 'equip'.
    """
    base = Path(stem).stem if stem else ""
    low = base.lower().replace(" ", "_").replace("-", "_")
    m = re.search(r"(?:^|[_\s])([abcd])(?:[_\s]|$)", low)
    if not m:
        m = re.search(r"report[_]?([abcd])\b", low)
    if not m:
        m = re.search(r"statement[_]?([abcd])\b", low)
    if not m:
        return None
    letter = m.group(1).upper()
    try:
        stmt = Statement(letter)
    except ValueError:
        return None
    cat = Category.EQUIPMENT if "equip" in low else Category.MEDICINE
    return stmt, cat


def _parse_statement_cell(val: Any) -> Optional[Statement]:
    if val is None:
        return None
    s = str(val).strip().upper()
    if len(s) == 1 and s in "ABCD":
        try:
            return Statement(s)
        except ValueError:
            return None
    if s.startswith("STATEMENT"):
        for ch in s:
            if ch in "ABCD":
                try:
                    return Statement(ch)
                except ValueError:
                    pass
    return None


def _parse_category_cell(val: Any, default: Category) -> Category:
    if val is None:
        return default
    low = str(val).lower()
    if "equip" in low:
        return Category.EQUIPMENT
    if "consum" in low:
        return Category.CONSUMABLES
    if "service" in low:
        return Category.SERVICES
    if "medicine" in low or ("med" in low and "edu" not in low):
        return Category.MEDICINE
    if "other" in low:
        return Category.OTHERS
    return default


def _find_header_row(ws, max_scan: int = 30) -> Tuple[Optional[int], List[str]]:
    """Locate header row: serial # column + item/description column."""
    limit = min(ws.max_row, max_scan)
    for r_idx in range(1, limit + 1):
        row = [str(c.value or "").strip() for c in ws[r_idx]]
        row_text = " ".join(row).lower()
        serial_ok = bool(
            re.search(r"(sr\.?\s*no|s\.?\s*no\.?|serial\s*no\.?|sl\.?\s*no\.?|^\s*#\s*$)", row_text)
            or any("sr" in c.lower() and "no" in c.lower() for c in row)
        )
        item_ok = bool(
            re.search(r"(name\s+of\s+item|item\s+name|item\s+description|description\s+of)", row_text)
            or ("item" in row_text and ("name" in row_text or "desc" in row_text))
        )
        if serial_ok and item_ok:
            return r_idx, row
    for r_idx in range(1, limit + 1):
        row = [str(c.value or "").strip() for c in ws[r_idx]]
        row_text = " ".join(row).lower()
        if (
            "name of item" in row_text
            or "category of items" in row_text
            or ("particulars" in row_text and ("amount" in row_text or "rs" in row_text or "₹" in row_text))
        ):
            return r_idx, row
    return None, []


# Default status per statement
STATEMENT_STATUS = {
    Statement.A: Status.PO_ISSUED,
    Statement.B: Status.TENDER_UNDER_PROCESS,
    Statement.C: Status.AWAITED_PUBLISH,
    Statement.D: Status.EXPIRED,
}


def classify_statement(status: Status, po_value: float) -> Statement:
    if status == Status.PO_ISSUED and po_value > 0:
        return Statement.A
    if status == Status.TENDER_UNDER_PROCESS:
        return Statement.B
    if status in (Status.AWAITED_PUBLISH, Status.RETENDER):
        return Statement.C
    if status in (Status.EXPIRED, Status.RETURNED, Status.CANCELLED, Status.CLOSED):
        return Statement.D
    return Statement.B


def compute_risk_level(
    status: Status,
    value_cr: float,
    days_pending: int,
    outstanding_cr: float,
) -> RiskLevel:
    # Rules applied in order, first match wins
    d_statuses = {Status.EXPIRED, Status.RETURNED, Status.CANCELLED, Status.CLOSED}
    backlog_statuses = {Status.AWAITED_PUBLISH, Status.RETENDER}
    if status in d_statuses and value_cr > 50 and days_pending > 90:
        return RiskLevel.CRITICAL
    if outstanding_cr > 10 or (status in backlog_statuses and days_pending > 60):
        return RiskLevel.HIGH
    if status in {Status.TENDER_UNDER_PROCESS, Status.AWAITED_PUBLISH, Status.RETENDER} and days_pending > 30:
        return RiskLevel.MEDIUM
    if status in d_statuses and value_cr > 10:
        return RiskLevel.HIGH
    if status in d_statuses:
        return RiskLevel.MEDIUM
    return RiskLevel.LOW


def compute_next_action(
    status: Status,
    value_cr: float,
    outstanding_cr: float,
    days_pending: int,
) -> str:
    if status == Status.AWAITED_PUBLISH:
        return "Publish tender immediately"
    if status == Status.RETENDER and value_cr > 10:
        return "Obtain retender approval — escalate to Dept Head"
    if status == Status.RETENDER:
        return "Initiate retender process"
    if status == Status.TENDER_UNDER_PROCESS and days_pending > 60:
        return "Expedite tender evaluation"
    if status == Status.TENDER_UNDER_PROCESS:
        return "Progress tender to PO stage"
    if outstanding_cr > 5 and status == Status.PO_ISSUED:
        return "Issue payment follow-up notice"
    if status == Status.EXPIRED and value_cr > 5:
        return "Review: retender or cancel — decision required"
    if status == Status.EXPIRED:
        return "Close expired tender after review"
    if status == Status.RETURNED:
        return "Address return remarks and resubmit"
    if outstanding_cr > 0 and status == Status.CANCELLED:
        return "Recover outstanding payment before closing"
    if status == Status.CANCELLED:
        return "Finalize cancellation"
    return "Monitor and update status"


def compute_priority_score(
    value_cr: float,
    risk_level: RiskLevel,
    days_pending: int,
    outstanding_cr: float,
) -> float:
    # Weighted: value 0.35, risk 0.30, ageing 0.20, outstanding/criticality 0.15
    risk_weight = {RiskLevel.CRITICAL: 100, RiskLevel.HIGH: 75,
                   RiskLevel.MEDIUM: 50, RiskLevel.LOW: 25}[risk_level]
    value_score = min(100, (value_cr / 100) * 100)  # scaled by 100 Cr
    ageing_score = min(100, (days_pending / 365) * 100)
    outstanding_score = min(100, (outstanding_cr / 50) * 100)
    return round(
        value_score * 0.35 + risk_weight * 0.30 + ageing_score * 0.20 + outstanding_score * 0.15,
        2,
    )


def estimate_days_pending(proposal_date: Any, statement: Statement) -> int:
    """Estimate days since proposal was received."""
    if proposal_date is None:
        return 30 if statement != Statement.A else 0
    try:
        if isinstance(proposal_date, datetime):
            delta = (datetime.now() - proposal_date).days
            return max(0, delta)
        s = str(proposal_date).strip()
        # Try common formats
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(s, fmt)
                return max(0, (datetime.now() - dt).days)
            except ValueError:
                continue
    except Exception:
        pass
    return 30


def fy_from_date(d: Any) -> str:
    """Derive financial year (Apr-Mar) from a date."""
    try:
        if isinstance(d, datetime):
            dt = d
        else:
            s = str(d).strip()
            dt = None
            for fmt in ("%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    break
                except ValueError:
                    continue
            if not dt:
                return "2024-25"
        # Fiscal year starts April
        year = dt.year
        if dt.month < 4:
            year -= 1
        return f"{year}-{str(year + 1)[2:]}"
    except Exception:
        return "2024-25"


def normalize_department(dept: Any) -> str:
    if not dept:
        return "Unknown"
    s = str(dept).strip()
    # Normalize common variations
    low = s.lower()
    if "public health" in low:
        return "Public Health Department"
    if "medical education" in low or "dmer" in low:
        return "Medical Education & Research"
    if "nhm" in low:
        return "NHM"
    if "sbtc" in low:
        return "SBTC"
    return s[:200]


def parse_workbook(path: str) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    """Parse the ABCD workbook -> list of procurement records. Returns (records, stats)."""
    if not Path(path).exists():
        return [], {"error": 1, "total": 0, "inserted": 0}

    wb = load_workbook(path, data_only=True)
    records: List[Dict[str, Any]] = []
    stats = {"total": 0, "inserted": 0, "skipped": 0, "errors": 0}
    stem = Path(path).stem
    filename_hint = _infer_stmt_cat_from_filename(stem)

    for sname in wb.sheetnames:
        stmt_cat = _lookup_sheet_name(sname)
        if not stmt_cat and filename_hint:
            st, cat = filename_hint
            sn = _normalize_sheet_title(sname)
            if "equip" in sn:
                cat = Category.EQUIPMENT
            elif "medicine" in sn and "equip" not in sn:
                cat = Category.MEDICINE
            stmt_cat = (st, cat)
        if not stmt_cat:
            continue
        sheet_statement, sheet_category = stmt_cat
        ws = wb[sname]

        header_row_idx, headers = _find_header_row(ws)
        if header_row_idx is None or not headers:
            continue

        # Map columns by matching names
        def find_col(patterns: List[str]) -> Optional[int]:
            for i, h in enumerate(headers):
                h_low = h.lower()
                if any(p.lower() in h_low for p in patterns):
                    return i
            return None

        col_sr = find_col(["sr", "s no", "serial", "sl no", "sl."])
        col_date = find_col(["date of proposal", "proposal date", "date"])
        col_dept = find_col(["received from", "department", "dept", "hod"])
        col_bureau = find_col(["bureau"])
        col_aa = find_col(["administrative approval", "aa date", "approval"])
        col_po_num = find_col(["po number", "po no", "purchase order"])
        col_item = find_col(
            [
                "name of item",
                "category of items",
                "item description",
                "description of item",
                "item name",
                "particulars",
            ]
        )
        col_value = find_col(
            [
                "total amount of purchase order",
                "total amount of proposed",
                "total amount",
                "amount in rs",
                "procurement value",
                "po value",
                "estimated cost",
            ]
        )
        col_budget = find_col(["budget source", "budget"])
        col_received = find_col(["amount received against", "received"])
        col_balance = find_col(["balance amount", "balance", "outstanding"])
        col_tender = find_col(["tender number", "tender no", "tender"])
        col_stmt = find_col(["statement", "abc", "stage", "classification"])
        col_cat = find_col(["category", "medicine / equipment", "item category", "type"])

        if col_item is None:
            continue

        # Iterate data rows
        for r_idx in range(header_row_idx + 1, ws.max_row + 1):
            row = [c.value for c in ws[r_idx]]
            if col_item >= len(row):
                continue
            item_raw = row[col_item]
            if not item_raw or not str(item_raw).strip():
                continue  # skip blank

            statement = sheet_statement
            category = sheet_category
            row_stmt = _parse_statement_cell(row[col_stmt]) if col_stmt is not None and col_stmt < len(row) else None
            if row_stmt is not None:
                statement = row_stmt
            if col_cat is not None and col_cat < len(row):
                category = _parse_category_cell(row[col_cat], sheet_category)

            default_status = STATEMENT_STATUS[statement]

            stats["total"] += 1
            try:
                value_raw = row[col_value] if col_value is not None and col_value < len(row) else 0
                procurement_value = parse_rupee_value(value_raw)
                if procurement_value <= 0 and statement != Statement.D:
                    # Statement D may have 0 (expired)
                    # but still keep — may have received amt
                    pass

                dept = normalize_department(row[col_dept] if col_dept is not None and col_dept < len(row) else "Public Health Department")
                bureau = str(row[col_bureau] or "").strip()[:200] if col_bureau is not None and col_bureau < len(row) else ""
                po_number = str(row[col_po_num] or "").strip()[:100] if col_po_num is not None and col_po_num < len(row) else ""
                tender_num = str(row[col_tender] or "").strip()[:100] if col_tender is not None and col_tender < len(row) else ""
                budget_src = str(row[col_budget] or "").strip()[:200] if col_budget is not None and col_budget < len(row) else ""
                proposal_date = row[col_date] if col_date is not None and col_date < len(row) else None
                approval_date = row[col_aa] if col_aa is not None and col_aa < len(row) else None
                received = parse_rupee_value(row[col_received]) if col_received is not None and col_received < len(row) else 0.0
                balance = parse_rupee_value(row[col_balance]) if col_balance is not None and col_balance < len(row) else 0.0

                # Statement A: po_value = procurement_value; paid = po_value - balance
                # Statement B/C/D: po_value = 0 (no PO yet)
                po_value = 0.0
                paid_amount = 0.0
                outstanding = 0.0
                if statement == Statement.A:
                    po_value = procurement_value
                    # received is amount received against proposal; balance is outstanding to pay
                    # but field semantics vary — use balance as outstanding if positive
                    paid_amount = max(0.0, po_value - balance) if balance > 0 else min(received, po_value)
                    outstanding = max(0.0, po_value - paid_amount)

                days_pending = estimate_days_pending(proposal_date, statement)

                status = default_status
                # For D, vary status based on sheet name heuristics
                if statement == Statement.D:
                    # Most D records are expired; no sub-classification available
                    status = Status.EXPIRED

                risk_level = compute_risk_level(status, procurement_value, days_pending, outstanding)
                action = compute_next_action(status, procurement_value, outstanding, days_pending)
                priority = compute_priority_score(procurement_value, risk_level, days_pending, outstanding)
                fy = fy_from_date(proposal_date)

                item_desc = str(item_raw).strip()[:500]
                record = {
                    "id": str(uuid.uuid4()),
                    "statement": statement.value,
                    "department": dept,
                    "bureau": bureau,
                    "category": category.value,
                    "item_description": item_desc,
                    "procurement_value": round(procurement_value, 4),
                    "po_value": round(po_value, 4),
                    "paid_amount": round(paid_amount, 4),
                    "outstanding_amount": round(outstanding, 4),
                    "current_status": status.value,
                    "risk_level": risk_level.value,
                    "action_required": action,
                    "priority_score": priority,
                    "days_pending": days_pending,
                    "financial_year": fy,
                    "budget_source": budget_src,
                    "po_number": po_number,
                    "tender_number": tender_num,
                    "proposal_date": str(proposal_date)[:30] if proposal_date else None,
                    "approval_date": str(approval_date)[:30] if approval_date else None,
                    "assigned_to": None,
                    "escalation_level": 0,
                    "due_date": None,
                    "remarks": None,
                    "data_source": Path(path).name,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
                records.append(record)
                stats["inserted"] += 1
            except Exception:
                stats["errors"] += 1
                continue

    return records, stats


def _norm_csv_key(k: str) -> str:
    return re.sub(r"\s+", "_", str(k or "").strip().lower())


def parse_csv_bytes(
    data: bytes,
    *,
    batch_id: Optional[str] = None,
    source_name: str = "upload.csv",
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Parse UTF-8 (with BOM) CSV into procurement records using flexible headers."""
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], {"total": 0, "inserted": 0, "errors": 1, "message": "empty_csv"}

    records: List[Dict[str, Any]] = []
    stats = {"total": 0, "inserted": 0, "errors": 0}
    upload_ts = datetime.now(timezone.utc).isoformat()

    for raw in reader:
        stats["total"] += 1
        row = {_norm_csv_key(k): (v.strip() if isinstance(v, str) else v) for k, v in raw.items() if k}
        try:
            item = row.get("item_description") or row.get("name_of_item") or row.get("item") or row.get("particulars")
            if not item:
                continue
            stmt = _parse_statement_cell(row.get("statement") or row.get("stage") or row.get("abc")) or Statement.A
            cat_default = Category.MEDICINE
            category = _parse_category_cell(row.get("category") or row.get("item_category"), cat_default)
            procurement_value = parse_rupee_value(row.get("procurement_value") or row.get("total_amount") or row.get("amount") or 0)
            dept = normalize_department(row.get("department") or row.get("received_from") or "Public Health Department")
            po_value = parse_rupee_value(row.get("po_value")) if row.get("po_value") else (procurement_value if stmt == Statement.A else 0.0)
            paid = parse_rupee_value(row.get("paid_amount") or row.get("received") or 0)
            outstanding = max(0.0, po_value - paid) if stmt == Statement.A else 0.0
            st_raw = str(row.get("current_status") or row.get("status") or "").strip()
            if st_raw:
                try:
                    status = Status(st_raw)
                except ValueError:
                    status = STATEMENT_STATUS[stmt]
            else:
                status = STATEMENT_STATUS[stmt]
            days_pending = int(float(row.get("days_pending") or 30))
            risk_level = compute_risk_level(status, procurement_value, days_pending, outstanding)
            action = compute_next_action(status, procurement_value, outstanding, days_pending)
            priority = compute_priority_score(procurement_value, risk_level, days_pending, outstanding)
            fy = row.get("financial_year") or fy_from_date(row.get("proposal_date"))

            rid = str(uuid.uuid4())
            record = {
                "id": rid,
                "record_id": rid,
                "batch_id": batch_id,
                "upload_date": upload_ts,
                "statement": stmt.value,
                "department": dept,
                "bureau": (row.get("bureau") or "")[:200],
                "category": category.value,
                "item_description": str(item)[:500],
                "procurement_value": round(procurement_value, 4),
                "po_value": round(po_value, 4),
                "paid_amount": round(paid, 4),
                "outstanding_amount": round(outstanding, 4),
                "current_status": status.value,
                "risk_level": risk_level.value,
                "action_required": action,
                "priority_score": priority,
                "days_pending": days_pending,
                "financial_year": fy if isinstance(fy, str) else str(fy),
                "budget_source": (row.get("budget_source") or "")[:200] or None,
                "po_number": (row.get("po_number") or "")[:100] or None,
                "tender_number": (row.get("tender_number") or "")[:100] or None,
                "proposal_date": str(row.get("proposal_date") or "")[:30] or None,
                "approval_date": str(row.get("approval_date") or "")[:30] or None,
                "assigned_to": None,
                "escalation_level": int(float(row.get("escalation_level") or 0)),
                "due_date": None,
                "remarks": (row.get("remarks") or "")[:2000] or None,
                "data_source": source_name,
                "created_at": upload_ts,
                "updated_at": upload_ts,
            }
            records.append(record)
            stats["inserted"] += 1
        except Exception:
            stats["errors"] += 1
    return records, stats


def data_quality_score(stats: Dict[str, Any], n_sample: int) -> float:
    """Heuristic 0–100 score from parse stats."""
    total = max(1, int(stats.get("total") or 0))
    err = int(stats.get("errors") or 0) + int(stats.get("skipped") or 0)
    ins = int(stats.get("inserted") or 0)
    ratio = ins / total
    penalty = min(40.0, (err / total) * 100)
    return max(0.0, min(100.0, round(ratio * 100 - penalty, 2)))
